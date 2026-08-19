from flask import Flask, jsonify, render_template_string
from openpyxl import load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "industrial_iot_data.xlsx"

HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>Industrial IoT Monitor</title>

    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f2f2f2;
            margin: 0;
            padding: 20px;
        }

        h1 {
            text-align: center;
        }

        .cards {
            display: flex;
            justify-content: center;
            gap: 20px;
            flex-wrap: wrap;
            margin-bottom: 25px;
        }

        .card {
            background: white;
            padding: 20px;
            width: 180px;
            text-align: center;
            border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
        }

        .value {
            font-size: 30px;
            font-weight: bold;
        }

        .chart-container {
            background: white;
            padding: 20px;
            margin-bottom: 25px;
            border-radius: 10px;
        }

        canvas {
            max-height: 350px;
        }
    </style>
</head>

<body>

<h1>INDUSTRIAL IoT MONITOR</h1>

<div class="cards">

    <div class="card">
        <h3>People Inside</h3>
        <div class="value" id="people">--</div>
    </div>

    <div class="card">
        <h3>Temperature</h3>
        <div class="value" id="temperature">-- °C</div>
    </div>

    <div class="card">
        <h3>Humidity</h3>
        <div class="value" id="humidity">-- %</div>
    </div>

</div>


<div class="chart-container">
    <h2>People Inside</h2>
    <canvas id="peopleChart"></canvas>
</div>


<div class="chart-container">
    <h2>Temperature</h2>
    <canvas id="temperatureChart"></canvas>
</div>


<div class="chart-container">
    <h2>Humidity</h2>
    <canvas id="humidityChart"></canvas>
</div>


<script>

let peopleChart;
let temperatureChart;
let humidityChart;


function createCharts() {

    peopleChart = new Chart(
        document.getElementById("peopleChart"),
        {
            type: "line",

            data: {
                labels: [],
                datasets: [{
                    label: "People Inside",
                    data: [],
                    tension: 0.2
                }]
            },

            options: {
                responsive: true,
                animation: false
            }
        }
    );


    temperatureChart = new Chart(
        document.getElementById("temperatureChart"),
        {
            type: "line",

            data: {
                labels: [],
                datasets: [{
                    label: "Temperature (°C)",
                    data: [],
                    tension: 0.2
                }]
            },

            options: {
                responsive: true,
                animation: false
            }
        }
    );


    humidityChart = new Chart(
        document.getElementById("humidityChart"),
        {
            type: "line",

            data: {
                labels: [],
                datasets: [{
                    label: "Humidity (%)",
                    data: [],
                    tension: 0.2
                }]
            },

            options: {
                responsive: true,
                animation: false
            }
        }
    );
}


async function updateDashboard() {

    try {

        const response = await fetch("/data");

        const data = await response.json();


        if (data.length === 0) {
            return;
        }


        const latest = data[data.length - 1];


        document.getElementById("people").innerText =
            latest.people;

        document.getElementById("temperature").innerText =
            latest.temperature + " °C";

        document.getElementById("humidity").innerText =
            latest.humidity + " %";


        const labels = data.map(x => x.time);

        const people = data.map(x => x.people);

        const temperatures = data.map(x => x.temperature);

        const humidity = data.map(x => x.humidity);


        peopleChart.data.labels = labels;
        peopleChart.data.datasets[0].data = people;
        peopleChart.update();


        temperatureChart.data.labels = labels;
        temperatureChart.data.datasets[0].data = temperatures;
        temperatureChart.update();


        humidityChart.data.labels = labels;
        humidityChart.data.datasets[0].data = humidity;
        humidityChart.update();

    }

    catch (error) {

        console.log("Dashboard error:", error);

    }
}


createCharts();

updateDashboard();


// Update every 2 seconds
setInterval(updateDashboard, 2000);

</script>

</body>
</html>
"""


@app.route("/")
def home():
    return render_template_string(HTML)


@app.route("/data")
def data():

    if not os.path.exists(EXCEL_FILE):
        return jsonify([])

    try:

        workbook = load_workbook(
            EXCEL_FILE,
            read_only=True,
            data_only=True
        )

        sheet = workbook.active

        rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if len(row) >= 4:

                time_value = row[0]
                people = row[1]
                temperature = row[2]
                humidity = row[3]

                if people is None:
                    continue

                rows.append({
                    "time": str(time_value),
                    "people": float(people),
                    "temperature": float(temperature),
                    "humidity": float(humidity)
                })

        workbook.close()

        # Show the most recent 100 readings
        rows = rows[-100:]

        return jsonify(rows)

    except Exception as e:

        print("Excel error:", e)

        return jsonify([])


if __name__ == "__main__":

    print("=================================")
    print(" INDUSTRIAL IoT DASHBOARD")
    print("=================================")
    print("Dashboard starting...")
    print("Open: http://127.0.0.1:5000")
    print()

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=False
    )