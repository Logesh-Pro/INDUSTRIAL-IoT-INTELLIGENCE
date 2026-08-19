from flask import Flask, jsonify, render_template_string, request
from openpyxl import load_workbook
import os
import requests
import statistics
from datetime import datetime

app = Flask(__name__)

# =========================================================
# SETTINGS
# =========================================================

EXCEL_FILE = "industrial_iot_data.xlsx"

# ESP32 address
ESP32_URL = "http://192.168.4.1/data"

# How many Excel readings to analyze
MAX_HISTORY = 500


# =========================================================
# EXCEL DATA
# =========================================================

def read_excel_data():

    if not os.path.exists(EXCEL_FILE):
        return []

    try:

        workbook = load_workbook(
            EXCEL_FILE,
            read_only=True,
            data_only=True
        )

        sheet = workbook.active

        rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if len(row) < 4:
                continue

            time_value = row[0]
            people = row[1]
            temperature = row[2]
            humidity = row[3]

            if people is None or temperature is None or humidity is None:
                continue

            try:

                rows.append({
                    "time": str(time_value),
                    "people": float(people),
                    "temperature": float(temperature),
                    "humidity": float(humidity)
                })

            except:
                continue

        workbook.close()

        return rows[-MAX_HISTORY:]

    except Exception as e:

        print("Excel error:", e)
        return []


# =========================================================
# LIVE ESP32 DATA
# =========================================================

def get_live_esp32():

    try:

        response = requests.get(
            ESP32_URL,
            timeout=1.5
        )

        text = response.text.strip()

        # -------------------------------------------------
        # JSON FORMAT
        # -------------------------------------------------

        try:

            data = response.json()

            if isinstance(data, dict):

                people = data.get("people")
                temperature = data.get("temperature")
                humidity = data.get("humidity")

                if people is not None:

                    return {
                        "online": True,
                        "people": float(people),
                        "temperature": float(temperature),
                        "humidity": float(humidity),
                        "time": datetime.now().strftime("%H:%M:%S")
                    }

        except:
            pass

        # -------------------------------------------------
        # CSV FORMAT
        # Example:
        # 13,30.2,67.4
        # -------------------------------------------------

        parts = text.split(",")

        if len(parts) >= 3:

            people = float(parts[0].strip())
            temperature = float(parts[1].strip())
            humidity = float(parts[2].strip())

            return {
                "online": True,
                "people": people,
                "temperature": temperature,
                "humidity": humidity,
                "time": datetime.now().strftime("%H:%M:%S")
            }

        return {
            "online": False
        }

    except Exception as e:

        return {
            "online": False,
            "error": str(e)
        }


# =========================================================
# STATISTICS
# =========================================================

def calculate_statistics(rows):

    if not rows:
        return {}

    people = [x["people"] for x in rows]
    temperatures = [x["temperature"] for x in rows]
    humidity = [x["humidity"] for x in rows]

    return {

        "average_people": round(statistics.mean(people), 2),
        "peak_people": max(people),
        "minimum_people": min(people),

        "average_temperature":
            round(statistics.mean(temperatures), 2),

        "maximum_temperature":
            round(max(temperatures), 2),

        "minimum_temperature":
            round(min(temperatures), 2),

        "average_humidity":
            round(statistics.mean(humidity), 2),

        "maximum_humidity":
            round(max(humidity), 2),

        "minimum_humidity":
            round(min(humidity), 2),

        "reading_count": len(rows)
    }


# =========================================================
# TREND
# =========================================================

def get_trend(values):

    if len(values) < 5:
        return "STABLE"

    recent = values[-5:]

    first = statistics.mean(recent[:2])
    last = statistics.mean(recent[-2:])

    difference = last - first

    if difference > 0.2:
        return "RISING"

    if difference < -0.2:
        return "FALLING"

    return "STABLE"


# =========================================================
# ANOMALY DETECTION
# =========================================================

def detect_anomalies(rows):

    if len(rows) < 10:
        return []

    anomalies = []

    temperatures = [x["temperature"] for x in rows]
    humidity = [x["humidity"] for x in rows]
    people = [x["people"] for x in rows]

    avg_temp = statistics.mean(temperatures)
    avg_humidity = statistics.mean(humidity)
    avg_people = statistics.mean(people)

    temp_std = statistics.pstdev(temperatures)
    humidity_std = statistics.pstdev(humidity)
    people_std = statistics.pstdev(people)

    latest = rows[-1]

    # Avoid zero standard deviation
    if temp_std > 0:
        if abs(latest["temperature"] - avg_temp) > 2 * temp_std:
            anomalies.append("Temperature is unusually different from normal.")

    if humidity_std > 0:
        if abs(latest["humidity"] - avg_humidity) > 2 * humidity_std:
            anomalies.append("Humidity is unusually different from normal.")

    if people_std > 0:
        if abs(latest["people"] - avg_people) > 2 * people_std:
            anomalies.append("Occupancy is unusually different from normal.")

    return anomalies


# =========================================================
# CORRELATION
# =========================================================

def correlation(x, y):

    if len(x) < 3:
        return 0

    try:

        mean_x = statistics.mean(x)
        mean_y = statistics.mean(y)

        numerator = sum(
            (a - mean_x) * (b - mean_y)
            for a, b in zip(x, y)
        )

        denominator_x = sum(
            (a - mean_x) ** 2
            for a in x
        )

        denominator_y = sum(
            (b - mean_y) ** 2
            for b in y
        )

        denominator = (
            denominator_x * denominator_y
        ) ** 0.5

        if denominator == 0:
            return 0

        return numerator / denominator

    except:
        return 0


# =========================================================
# SIMPLE PREDICTION
# =========================================================

def predict_next(values):

    if len(values) < 5:
        return None

    recent = values[-5:]

    differences = []

    for i in range(1, len(recent)):
        differences.append(
            recent[i] - recent[i - 1]
        )

    average_change = statistics.mean(differences)

    prediction = recent[-1] + average_change

    return round(prediction, 2)


# =========================================================
# SMART INSIGHTS
# =========================================================

def generate_insight(rows, live):

    if not rows:
        return "Waiting for historical data."

    stats = calculate_statistics(rows)

    latest = live if live.get("online") else rows[-1]

    people_trend = get_trend(
        [x["people"] for x in rows]
    )

    temp_trend = get_trend(
        [x["temperature"] for x in rows]
    )

    humidity_trend = get_trend(
        [x["humidity"] for x in rows]
    )

    if latest["people"] >= stats["peak_people"]:
        occupancy_message = (
            "Occupancy is currently at or near the recorded peak."
        )

    elif people_trend == "RISING":
        occupancy_message = (
            "Occupancy is currently increasing."
        )

    elif people_trend == "FALLING":
        occupancy_message = (
            "Occupancy is currently decreasing."
        )

    else:
        occupancy_message = (
            "Occupancy is relatively stable."
        )

    if latest["temperature"] >= 32:
        environment = "High-temperature condition detected."

    elif latest["temperature"] >= 28:
        environment = "Warm environment."

    else:
        environment = "Temperature is within the lower operating range."

    return occupancy_message + " " + environment


# =========================================================
# EVENTS
# =========================================================

def generate_events(rows):

    if len(rows) < 2:
        return []

    events = []

    for i in range(1, len(rows)):

        previous = rows[i - 1]
        current = rows[i]

        people_change = (
            current["people"] -
            previous["people"]
        )

        temp_change = (
            current["temperature"] -
            previous["temperature"]
        )

        humidity_change = (
            current["humidity"] -
            previous["humidity"]
        )

        if people_change >= 2:

            events.append({
                "time": current["time"],
                "type": "OCCUPANCY SURGE",
                "message":
                    f"Occupancy increased by {people_change:.0f}."
            })

        elif people_change <= -2:

            events.append({
                "time": current["time"],
                "type": "OCCUPANCY DROP",
                "message":
                    f"Occupancy decreased by {abs(people_change):.0f}."
            })

        if temp_change >= 0.5:

            events.append({
                "time": current["time"],
                "type": "TEMPERATURE RISE",
                "message":
                    f"Temperature increased by {temp_change:.1f} °C."
            })

        elif temp_change <= -0.5:

            events.append({
                "time": current["time"],
                "type": "TEMPERATURE DROP",
                "message":
                    f"Temperature decreased by {abs(temp_change):.1f} °C."
            })

        if humidity_change >= 3:

            events.append({
                "time": current["time"],
                "type": "HUMIDITY RISE",
                "message":
                    f"Humidity increased by {humidity_change:.1f}%."
            })

    return events[-20:]


# =========================================================
# API
# =========================================================

@app.route("/api/dashboard")
def dashboard_data():

    rows = read_excel_data()

    live = get_live_esp32()

    stats = calculate_statistics(rows)

    anomalies = detect_anomalies(rows)

    events = generate_events(rows)

    if rows:

        people_values = [
            x["people"] for x in rows
        ]

        temp_values = [
            x["temperature"] for x in rows
        ]

        humidity_values = [
            x["humidity"] for x in rows
        ]

        people_temp_corr = correlation(
            people_values,
            temp_values
        )

        people_humidity_corr = correlation(
            people_values,
            humidity_values
        )

        predicted_people = predict_next(
            people_values
        )

        predicted_temperature = predict_next(
            temp_values
        )

        predicted_humidity = predict_next(
            humidity_values
        )

    else:

        people_temp_corr = 0
        people_humidity_corr = 0

        predicted_people = None
        predicted_temperature = None
        predicted_humidity = None

    return jsonify({

        "live": live,

        "statistics": stats,

        "trend": {

            "people": get_trend(
                [x["people"] for x in rows]
            ) if rows else "STABLE",

            "temperature": get_trend(
                [x["temperature"] for x in rows]
            ) if rows else "STABLE",

            "humidity": get_trend(
                [x["humidity"] for x in rows]
            ) if rows else "STABLE"
        },

        "prediction": {

            "people": predicted_people,

            "temperature":
                predicted_temperature,

            "humidity":
                predicted_humidity
        },

        "correlation": {

            "people_temperature":
                round(people_temp_corr, 2),

            "people_humidity":
                round(people_humidity_corr, 2)
        },

        "anomalies": anomalies,

        "events": events,

        "insight":
            generate_insight(rows, live),

        "history": rows[-100:]
    })


# =========================================================
# WEBSITE
# =========================================================

HTML = """
<!DOCTYPE html>

<html>

<head>

<meta charset="UTF-8">

<title>Industrial IoT Intelligence Dashboard</title>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

<style>
/* =========================================================
   INDUSTRIAL IoT DASHBOARD — MODERN SCADA STYLE
   ========================================================= */

:root {
    --bg: #07111f;
    --bg2: #0b1728;
    --panel: #101f32;
    --panel2: #14263d;
    --border: #243b55;

    --text: #edf6ff;
    --muted: #8da4ba;

    --cyan: #00d9ff;
    --blue: #3498ff;
    --green: #20e6a2;
    --yellow: #ffd166;
    --orange: #ff9f43;
    --red: #ff5263;
    --purple: #a970ff;

    --shadow:
        0 10px 30px rgba(0, 0, 0, 0.28);

    --radius: 16px;
}


/* =========================================================
   RESET
   ========================================================= */

* {
    box-sizing: border-box;
}

html {
    scroll-behavior: smooth;
}

body {
    margin: 0;

    font-family:
        Inter,
        "Segoe UI",
        Roboto,
        Arial,
        sans-serif;

    background:
        radial-gradient(
            circle at 20% 0%,
            rgba(0, 217, 255, 0.08),
            transparent 30%
        ),
        radial-gradient(
            circle at 100% 20%,
            rgba(52, 152, 255, 0.08),
            transparent 30%
        ),
        var(--bg);

    color: var(--text);

    min-height: 100vh;
}


/* =========================================================
   HEADER
   ========================================================= */

header {
    position: relative;

    padding: 32px 20px;

    text-align: center;

    background:
        linear-gradient(
            135deg,
            #081522,
            #10243b
        );

    border-bottom:
        1px solid var(--border);

    box-shadow:
        0 5px 25px rgba(0, 0, 0, 0.35);

    overflow: hidden;
}

header::before {
    content: "";

    position: absolute;

    inset: 0;

    background:
        linear-gradient(
            90deg,
            transparent 0%,
            rgba(0, 217, 255, 0.06) 50%,
            transparent 100%
        );

    animation:
        scan 5s linear infinite;
}

@keyframes scan {
    from {
        transform: translateX(-100%);
    }

    to {
        transform: translateX(100%);
    }
}

header h1 {
    position: relative;

    margin: 0;

    font-size: clamp(24px, 4vw, 38px);

    font-weight: 800;

    letter-spacing: 2px;

    color: white;

    text-shadow:
        0 0 18px rgba(0, 217, 255, 0.35);
}

header p {
    position: relative;

    margin: 10px 0 0;

    color: var(--muted);

    font-size: 14px;

    letter-spacing: 1px;
}


/* =========================================================
   MAIN CONTAINER
   ========================================================= */

.container {
    width: min(1500px, 94%);

    margin: auto;

    padding: 28px 0 50px;
}


/* =========================================================
   SECTION / PANEL
   ========================================================= */

.section {
    position: relative;

    background:
        linear-gradient(
            145deg,
            rgba(16, 31, 50, 0.97),
            rgba(11, 23, 40, 0.97)
        );

    border:
        1px solid var(--border);

    border-radius: var(--radius);

    padding: 24px;

    margin-bottom: 22px;

    box-shadow: var(--shadow);

    overflow: hidden;
}

.section::before {
    content: "";

    position: absolute;

    top: 0;
    left: 0;

    width: 100%;
    height: 2px;

    background:
        linear-gradient(
            90deg,
            transparent,
            var(--cyan),
            transparent
        );

    opacity: 0.55;
}

.section h2 {
    margin: 0 0 22px;

    padding-bottom: 14px;

    border-bottom:
        1px solid var(--border);

    color: white;

    font-size: 18px;

    font-weight: 700;

    letter-spacing: 0.5px;
}


/* =========================================================
   CARD GRID
   ========================================================= */

.cards {
    display: grid;

    grid-template-columns:
        repeat(auto-fit, minmax(190px, 1fr));

    gap: 16px;
}


/* =========================================================
   CARDS
   ========================================================= */

.card {
    position: relative;

    background:
        linear-gradient(
            145deg,
            #13263c,
            #0d1d30
        );

    border:
        1px solid var(--border);

    border-radius: 14px;

    padding: 20px;

    text-align: center;

    box-shadow:
        0 6px 18px rgba(0, 0, 0, 0.2);

    transition:
        transform 0.2s ease,
        border-color 0.2s ease,
        box-shadow 0.2s ease;
}

.card:hover {
    transform: translateY(-3px);

    border-color:
        rgba(0, 217, 255, 0.5);

    box-shadow:
        0 10px 30px rgba(0, 217, 255, 0.08);
}

.card h3 {
    margin: 0;

    color: var(--muted);

    font-size: 12px;

    text-transform: uppercase;

    letter-spacing: 1px;
}

.big {
    margin: 13px 0 8px;

    font-size: 32px;

    line-height: 1;

    font-weight: 800;

    color: white;

    text-shadow:
        0 0 15px rgba(0, 217, 255, 0.15);
}


/* =========================================================
   STATUS
   ========================================================= */

.status {
    display: inline-flex;

    align-items: center;

    gap: 8px;

    padding: 7px 13px;

    border-radius: 30px;

    font-size: 11px;

    font-weight: 800;

    letter-spacing: 1px;
}

.status::before {
    content: "";

    width: 7px;
    height: 7px;

    border-radius: 50%;
}

.online {
    background:
        rgba(32, 230, 162, 0.12);

    color: var(--green);

    border:
        1px solid rgba(32, 230, 162, 0.25);
}

.online::before {
    background: var(--green);

    box-shadow:
        0 0 10px var(--green);

    animation:
        pulse 1.5s infinite;
}

.offline {
    background:
        rgba(255, 82, 99, 0.12);

    color: var(--red);

    border:
        1px solid rgba(255, 82, 99, 0.25);
}

.offline::before {
    background: var(--red);
}

@keyframes pulse {
    0%,
    100% {
        opacity: 1;
        transform: scale(1);
    }

    50% {
        opacity: 0.45;
        transform: scale(0.75);
    }
}


/* =========================================================
   TWO COLUMN GRID
   ========================================================= */

.grid2 {
    display: grid;

    grid-template-columns:
        repeat(auto-fit, minmax(320px, 1fr));

    gap: 20px;
}


/* =========================================================
   CHARTS
   ========================================================= */

.chart-box {
    height: 350px;

    padding: 15px;

    background:
        rgba(7, 17, 31, 0.45);

    border:
        1px solid var(--border);

    border-radius: 12px;
}


/* =========================================================
   INSIGHT ENGINE
   ========================================================= */

.insight {
    position: relative;

    padding: 20px 22px;

    border-radius: 12px;

    background:
        linear-gradient(
            135deg,
            rgba(0, 217, 255, 0.09),
            rgba(52, 152, 255, 0.05)
        );

    border:
        1px solid rgba(0, 217, 255, 0.18);

    color: #dcefff;

    font-size: 16px;

    line-height: 1.7;
}

.insight::before {
    content: "AI";

    position: absolute;

    top: 12px;
    right: 14px;

    font-size: 10px;

    font-weight: 800;

    color: var(--cyan);

    opacity: 0.7;

    letter-spacing: 2px;
}


/* =========================================================
   ALERTS
   ========================================================= */

.alert,
.good {
    padding: 15px 18px;

    margin-bottom: 10px;

    border-radius: 10px;

    font-size: 14px;

    line-height: 1.5;
}

.alert {
    background:
        rgba(255, 82, 99, 0.08);

    border-left:
        4px solid var(--red);

    color: #ffd9dd;
}

.good {
    background:
        rgba(32, 230, 162, 0.07);

    border-left:
        4px solid var(--green);

    color: #c9ffe9;
}


/* =========================================================
   DIGITAL TWIN
   ========================================================= */

.digital-room {
    position: relative;

    min-height: 360px;

    padding: 28px;

    border-radius: 18px;

    background:
        linear-gradient(
            135deg,
            #101d2c,
            #172c42
        );

    border:
        1px solid #29445f;

    color: white;

    overflow: hidden;

    box-shadow:
        inset 0 0 80px rgba(0, 217, 255, 0.03);
}

.digital-room::before {
    content: "";

    position: absolute;

    inset: 0;

    background-image:
        linear-gradient(
            rgba(255,255,255,0.025) 1px,
            transparent 1px
        ),
        linear-gradient(
            90deg,
            rgba(255,255,255,0.025) 1px,
            transparent 1px
        );

    background-size: 35px 35px;

    pointer-events: none;
}

.room-title {
    position: relative;

    text-align: center;

    margin-bottom: 20px;

    color: var(--cyan);

    font-size: 18px;

    font-weight: 700;

    letter-spacing: 1px;

    text-transform: uppercase;
}

.people-area {
    position: relative;

    display: flex;

    flex-wrap: wrap;

    align-items: center;

    justify-content: center;

    gap: 8px;

    min-height: 110px;

    padding: 20px;

    border-radius: 12px;

    background:
        rgba(0, 0, 0, 0.16);

    border:
        1px solid rgba(255,255,255,0.05);
}

.person {
    display: inline-block;

    font-size: 27px;

    filter:
        drop-shadow(
            0 0 7px
            rgba(0, 217, 255, 0.4)
        );

    animation:
        personAppear 0.3s ease;
}

@keyframes personAppear {
    from {
        opacity: 0;
        transform: scale(0.5);
    }

    to {
        opacity: 1;
        transform: scale(1);
    }
}

.room-info {
    position: relative;

    display: flex;

    justify-content: space-around;

    flex-wrap: wrap;

    gap: 15px;

    margin-top: 22px;
}

.room-value {
    min-width: 130px;

    text-align: center;

    padding: 12px 20px;
}

.room-value strong {
    display: block;

    margin-top: 6px;

    font-size: 25px;

    color: white;
}


/* =========================================================
   OCCUPANCY PROGRESS
   ========================================================= */

.progress {
    position: relative;

    height: 10px;

    margin-top: 20px;

    background:
        #07111f;

    border-radius: 20px;

    overflow: hidden;

    border:
        1px solid #263d53;
}

.progress-bar {
    height: 100%;

    width: 0%;

    border-radius: inherit;

    background:
        linear-gradient(
            90deg,
            var(--green),
            var(--cyan),
            var(--blue)
        );

    box-shadow:
        0 0 14px
        rgba(0, 217, 255, 0.6);

    transition:
        width 0.6s ease;
}


/* =========================================================
   TIMELINE
   ========================================================= */

.timeline {
    border-left:
        2px solid #2b587e;

    padding-left: 22px;
}

.event {
    position: relative;

    padding: 15px 17px;

    margin-bottom: 12px;

    background:
        rgba(255, 255, 255, 0.025);

    border:
        1px solid var(--border);

    border-radius: 10px;

    color: #c9d8e7;
}

.event::before {
    content: "";

    position: absolute;

    left: -30px;

    top: 20px;

    width: 10px;
    height: 10px;

    border-radius: 50%;

    background: var(--cyan);

    box-shadow:
        0 0 10px var(--cyan);
}

.event strong {
    color: var(--cyan);

    font-size: 12px;

    letter-spacing: 0.5px;
}


/* =========================================================
   TABLE
   ========================================================= */

table {
    width: 100%;

    border-collapse: separate;

    border-spacing: 0;

    overflow: hidden;

    border:
        1px solid var(--border);

    border-radius: 12px;
}

th,
td {
    padding: 13px 12px;

    text-align: center;

    border-bottom:
        1px solid rgba(255,255,255,0.06);
}

th {
    background:
        #0b1929;

    color: var(--cyan);

    font-size: 11px;

    text-transform: uppercase;

    letter-spacing: 1px;
}

td {
    color: #c6d6e5;

    font-size: 13px;
}

tbody tr {
    transition:
        background 0.2s ease;
}

tbody tr:hover {
    background:
        rgba(0, 217, 255, 0.04);
}

tbody tr:last-child td {
    border-bottom: none;
}


/* =========================================================
   BUTTONS
   ========================================================= */

button {
    border: 1px solid rgba(0, 217, 255, 0.3);

    padding: 11px 18px;

    border-radius: 9px;

    cursor: pointer;

    background:
        linear-gradient(
            135deg,
            #1677a5,
            #145b8c
        );

    color: white;

    font-weight: 700;

    font-size: 12px;

    letter-spacing: 0.5px;

    margin: 5px;

    transition:
        transform 0.2s ease,
        box-shadow 0.2s ease,
        background 0.2s ease;
}

button:hover {
    transform: translateY(-2px);

    background:
        linear-gradient(
            135deg,
            #1b91c5,
            #176da3
        );

    box-shadow:
        0 7px 20px
        rgba(0, 217, 255, 0.18);
}

button:active {
    transform: translateY(0);
}


/* =========================================================
   EXPERIMENT RESULT
   ========================================================= */

.experiment-result {
    background:
        linear-gradient(
            135deg,
            rgba(169, 112, 255, 0.09),
            rgba(52, 152, 255, 0.05)
        );

    border:
        1px solid rgba(169, 112, 255, 0.2);

    border-left:
        4px solid var(--purple);

    padding: 20px;

    border-radius: 12px;

    margin-top: 15px;

    display: none;

    color: #dfd3f8;

    line-height: 1.6;
}


/* =========================================================
   SMALL TEXT
   ========================================================= */

.small {
    color: var(--muted);

    font-size: 12px;

    line-height: 1.5;
}


/* =========================================================
   FOOTER
   ========================================================= */

footer {
    text-align: center;

    padding: 28px 20px;

    color: #61788e;

    font-size: 12px;

    letter-spacing: 0.5px;

    border-top:
        1px solid rgba(255,255,255,0.04);
}


/* =========================================================
   SCROLLBAR
   ========================================================= */

::-webkit-scrollbar {
    width: 9px;
}

::-webkit-scrollbar-track {
    background: #07111f;
}

::-webkit-scrollbar-thumb {
    background: #24415c;

    border-radius: 10px;
}

::-webkit-scrollbar-thumb:hover {
    background: #32617f;
}


/* =========================================================
   MOBILE
   ========================================================= */

@media (max-width: 700px) {

    .container {
        width: 96%;

        padding-top: 15px;
    }

    .section {
        padding: 16px;

        border-radius: 12px;
    }

    .section h2 {
        font-size: 15px;
    }

    .cards {
        grid-template-columns:
            repeat(2, 1fr);

        gap: 10px;
    }

    .card {
        padding: 15px 10px;
    }

    .big {
        font-size: 24px;
    }

    .grid2 {
        grid-template-columns: 1fr;
    }

    .chart-box {
        height: 280px;
    }

    .digital-room {
        padding: 18px;

        min-height: 330px;
    }

    .room-info {
        gap: 5px;
    }

    .room-value {
        min-width: 90px;

        padding: 8px;
    }

    .room-value strong {
        font-size: 20px;
    }

    table {
        font-size: 11px;
    }

    th,
    td {
        padding: 9px 5px;
    }

    header h1 {
        font-size: 23px;
    }
}


/* =========================================================
   VERY SMALL SCREENS
   ========================================================= */

@media (max-width: 420px) {

    .cards {
        grid-template-columns: 1fr;
    }

    .big {
        font-size: 28px;
    }

    .people-area {
        gap: 5px;
    }

    .person {
        font-size: 22px;
    }
}

</style>

</head>


<body>


<header>

<h1>INDUSTRIAL IoT INTELLIGENCE</h1>

<p>
Real-Time Environmental & Occupancy Analytics
</p>

</header>


<div class="container">


<!-- ================================================= -->
<!-- LIVE -->
<!-- ================================================= -->

<div class="section">

<h2>🔴 LIVE SYSTEM</h2>

<div class="cards">


<div class="card">

<h3>People Inside</h3>

<div
    class="big"
    id="livePeople">
--
</div>

<div id="occupancyTrend">
--
</div>

</div>


<div class="card">

<h3>Temperature</h3>

<div
    class="big"
    id="liveTemperature">
--
</div>

<div id="temperatureTrend">
--
</div>

</div>


<div class="card">

<h3>Humidity</h3>

<div
    class="big"
    id="liveHumidity">
--
</div>

<div id="humidityTrend">
--
</div>

</div>


<div class="card">

<h3>ESP32</h3>

<div
    class="big"
    id="connection">
--
</div>

<div
    class="small"
    id="liveTime">
--
</div>

</div>


</div>

</div>


<!-- ================================================= -->
<!-- DIGITAL TWIN -->
<!-- ================================================= -->

<div class="section">

<h2>🏭 DIGITAL ROOM</h2>

<div class="digital-room">

<div class="room-title">
Industrial Monitoring Area
</div>


<div
    class="people-area"
    id="peopleArea">
</div>


<div class="room-info">

<div class="room-value">

People

<strong id="roomPeople">0</strong>

</div>


<div class="room-value">

Temperature

<strong id="roomTemperature">
-- °C
</strong>

</div>


<div class="room-value">

Humidity

<strong id="roomHumidity">
-- %
</strong>

</div>

</div>


<div class="progress">

<div
    class="progress-bar"
    id="occupancyBar">
</div>

</div>


<div
    style="
    text-align:center;
    margin-top:8px;
    "
>

Occupancy Level

</div>


</div>

</div>


<!-- ================================================= -->
<!-- SMART INSIGHT -->
<!-- ================================================= -->

<div class="section">

<h2>🧠 SMART INSIGHT ENGINE</h2>

<div
    class="insight"
    id="insight">
Analyzing environment...
</div>

</div>


<!-- ================================================= -->
<!-- ANOMALIES -->
<!-- ================================================= -->

<div class="section">

<h2>🚨 ANOMALY DETECTION</h2>

<div id="anomalies">

Checking data...

</div>

</div>


<!-- ================================================= -->
<!-- PREDICTION -->
<!-- ================================================= -->

<div class="section">

<h2>🔮 SHORT-TERM TREND PREDICTION</h2>

<div class="cards">


<div class="card">

<h3>Expected People</h3>

<div
    class="big"
    id="predPeople">
--
</div>

</div>


<div class="card">

<h3>Expected Temperature</h3>

<div
    class="big"
    id="predTemperature">
--
</div>

</div>


<div class="card">

<h3>Expected Humidity</h3>

<div
    class="big"
    id="predHumidity">
--
</div>

</div>


</div>

<p class="small">

Prediction is based on the recent trend in collected sensor readings.

</p>

</div>


<!-- ================================================= -->
<!-- BEHAVIOR -->
<!-- ================================================= -->

<div class="section">

<h2>👥 OCCUPANCY BEHAVIOR</h2>

<div class="cards">


<div class="card">

<h3>Average People</h3>

<div
    class="big"
    id="avgPeople">
--
</div>

</div>


<div class="card">

<h3>Peak Occupancy</h3>

<div
    class="big"
    id="peakPeople">
--
</div>

</div>


<div class="card">

<h3>Minimum Occupancy</h3>

<div
    class="big"
    id="minPeople">
--
</div>

</div>


<div class="card">

<h3>Total Readings</h3>

<div
    class="big"
    id="readingCount">
--
</div>

</div>


</div>

</div>


<!-- ================================================= -->
<!-- ENVIRONMENT -->
<!-- ================================================= -->

<div class="section">

<h2>🌡️ ENVIRONMENTAL ANALYTICS</h2>

<div class="cards">


<div class="card">

<h3>Average Temperature</h3>

<div
    class="big"
    id="avgTemp">
--
</div>

</div>


<div class="card">

<h3>Maximum Temperature</h3>

<div
    class="big"
    id="maxTemp">
--
</div>

</div>


<div class="card">

<h3>Minimum Temperature</h3>

<div
    class="big"
    id="minTemp">
--
</div>

</div>


<div class="card">

<h3>Average Humidity</h3>

<div
    class="big"
    id="avgHumidity">
--
</div>

</div>


</div>

</div>


<!-- ================================================= -->
<!-- CORRELATION -->
<!-- ================================================= -->

<div class="section">

<h2>🔬 PEOPLE vs ENVIRONMENT</h2>

<div class="grid2">


<div class="card">

<h3>People ↔ Temperature</h3>

<div
    class="big"
    id="peopleTempCorrelation">
--
</div>

<p id="peopleTempText">
Analyzing...
</p>

</div>


<div class="card">

<h3>People ↔ Humidity</h3>

<div
    class="big"
    id="peopleHumidityCorrelation">
--
</div>

<p id="peopleHumidityText">
Analyzing...
</p>

</div>


</div>

</div>


<!-- ================================================= -->
<!-- CHARTS -->
<!-- ================================================= -->

<div class="section">

<h2>📊 SENSOR HISTORY</h2>

<div class="grid2">


<div class="chart-box">

<canvas id="peopleChart"></canvas>

</div>


<div class="chart-box">

<canvas id="temperatureChart"></canvas>

</div>


<div class="chart-box">

<canvas id="humidityChart"></canvas>

</div>


</div>

</div>


<!-- ================================================= -->
<!-- EVENT TIMELINE -->
<!-- ================================================= -->

<div class="section">

<h2>🕵️ EVENT TIMELINE</h2>

<div
    class="timeline"
    id="timeline">

No events detected.

</div>

</div>


<!-- ================================================= -->
<!-- EXPERIMENT -->
<!-- ================================================= -->

<div class="section">

<h2>🧪 EXPERIMENT MODE</h2>

<p>

Use this mode when you want to perform a real-world experiment,
such as changing occupancy and observing environmental response.

</p>


<button onclick="startExperiment()">
START EXPERIMENT
</button>


<button onclick="finishExperiment()">
FINISH EXPERIMENT
</button>


<div
    class="experiment-result"
    id="experimentResult">
</div>

</div>


<!-- ================================================= -->
<!-- RECENT READINGS -->
<!-- ================================================= -->

<div class="section">

<h2>📋 RECENT SENSOR READINGS</h2>

<table>

<thead>

<tr>

<th>Time</th>

<th>People</th>

<th>Temperature</th>

<th>Humidity</th>

</tr>

</thead>

<tbody id="recentTable">

</tbody>

</table>

</div>


</div>


<footer>

Industrial IoT Monitoring & Analytics System

</footer>


<script>


let peopleChart;

let temperatureChart;

let humidityChart;

let experimentStart = null;


function makeCharts() {


peopleChart = new Chart(

document.getElementById("peopleChart"),

{

type: "line",

data: {

labels: [],

datasets: [{

label: "People Inside",

data: [],

tension: 0.25

}]

},

options: {

responsive: true,

animation: false

}

});


temperatureChart = new Chart(

document.getElementById("temperatureChart"),

{

type: "line",

data: {

labels: [],

datasets: [{

label: "Temperature °C",

data: [],

tension: 0.25

}]

},

options: {

responsive: true,

animation: false

}

});


humidityChart = new Chart(

document.getElementById("humidityChart"),

{

type: "line",

data: {

labels: [],

datasets: [{

label: "Humidity %",

data: [],

tension: 0.25

}]

},

options: {

responsive: true,

animation: false

}

});

}


function trendText(value) {

if (value === "RISING")
return "↑ Rising";

if (value === "FALLING")
return "↓ Falling";

return "→ Stable";

}


function correlationText(value) {

if (value > 0.6)
return "Strong positive relationship.";

if (value > 0.3)
return "Moderate positive relationship.";

if (value < -0.6)
return "Strong negative relationship.";

if (value < -0.3)
return "Moderate negative relationship.";

return "Weak relationship.";

}


async function updateDashboard() {


try {


const response =
await fetch("/api/dashboard");


const data =
await response.json();


const live =
data.live;


const stats =
data.statistics;


const trend =
data.trend;


const prediction =
data.prediction;


const correlation =
data.correlation;


<!-- LIVE -->


if (live.online) {


document.getElementById(
"livePeople"
).innerText =
live.people;


document.getElementById(
"liveTemperature"
).innerText =
live.temperature.toFixed(1) + " °C";


document.getElementById(
"liveHumidity"
).innerText =
live.humidity.toFixed(1) + " %";


document.getElementById(
"connection"
).innerHTML =
'<span class="status online">ONLINE</span>';


document.getElementById(
"liveTime"
).innerText =
"Updated " + live.time;


document.getElementById(
"roomPeople"
).innerText =
live.people;


document.getElementById(
"roomTemperature"
).innerText =
live.temperature.toFixed(1) + " °C";


document.getElementById(
"roomHumidity"
).innerText =
live.humidity.toFixed(1) + " %";


let peopleIcons = "";

let displayPeople =
Math.min(live.people, 30);


for (
let i = 0;
i < displayPeople;
i++
) {

peopleIcons +=
'<span class="person">👤</span>';

}


document.getElementById(
"peopleArea"
).innerHTML =
peopleIcons;


let occupancyPercent =
Math.min(
(live.people / 20) * 100,
100
);


document.getElementById(
"occupancyBar"
).style.width =
occupancyPercent + "%";


} else {


document.getElementById(
"connection"
).innerHTML =
'<span class="status offline">OFFLINE</span>';


}


<!-- TRENDS -->


document.getElementById(
"occupancyTrend"
).innerText =
trendText(trend.people);


document.getElementById(
"temperatureTrend"
).innerText =
trendText(trend.temperature);


document.getElementById(
"humidityTrend"
).innerText =
trendText(trend.humidity);


<!-- INSIGHT -->


document.getElementById(
"insight"
).innerText =
data.insight;


<!-- STATISTICS -->


if (stats.reading_count) {


document.getElementById(
"avgPeople"
).innerText =
stats.average_people;


document.getElementById(
"peakPeople"
).innerText =
stats.peak_people;


document.getElementById(
"minPeople"
).innerText =
stats.minimum_people;


document.getElementById(
"readingCount"
).innerText =
stats.reading_count;


document.getElementById(
"avgTemp"
).innerText =
stats.average_temperature +
" °C";


document.getElementById(
"maxTemp"
).innerText =
stats.maximum_temperature +
" °C";


document.getElementById(
"minTemp"
).innerText =
stats.minimum_temperature +
" °C";


document.getElementById(
"avgHumidity"
).innerText =
stats.average_humidity +
" %";


}


<!-- PREDICTION -->


document.getElementById(
"predPeople"
).innerText =
prediction.people ?? "--";


document.getElementById(
"predTemperature"
).innerText =
prediction.temperature !== null
?
prediction.temperature + " °C"
:
"--";


document.getElementById(
"predHumidity"
).innerText =
prediction.humidity !== null
?
prediction.humidity + " %"
:
"--";


<!-- CORRELATION -->


document.getElementById(
"peopleTempCorrelation"
).innerText =
correlation.people_temperature;


document.getElementById(
"peopleHumidityCorrelation"
).innerText =
correlation.people_humidity;


document.getElementById(
"peopleTempText"
).innerText =
correlationText(
correlation.people_temperature
);


document.getElementById(
"peopleHumidityText"
).innerText =
correlationText(
correlation.people_humidity
);


<!-- ANOMALIES -->


const anomalyBox =
document.getElementById(
"anomalies"
);


if (
data.anomalies.length === 0
) {


anomalyBox.innerHTML =
'<div class="good">' +
'🟢 No significant anomalies detected.' +
'</div>';


} else {


anomalyBox.innerHTML =
data.anomalies.map(
x =>
'<div class="alert">⚠️ ' +
x +
'</div>'
).join("");


}


<!-- CHARTS -->


const history =
data.history;


const labels =
history.map(x => x.time);


peopleChart.data.labels =
labels;


peopleChart.data.datasets[0].data =
history.map(x => x.people);


peopleChart.update();


temperatureChart.data.labels =
labels;


temperatureChart.data.datasets[0].data =
history.map(x => x.temperature);


temperatureChart.update();


humidityChart.data.labels =
labels;


humidityChart.data.datasets[0].data =
history.map(x => x.humidity);


humidityChart.update();


<!-- EVENTS -->


const timeline =
document.getElementById(
"timeline"
);


if (
data.events.length === 0
) {


timeline.innerHTML =
"No significant events detected.";

} else {


timeline.innerHTML =
data.events.map(
event =>

'<div class="event">' +

'<strong>' +
event.time +
'</strong> — ' +

event.type +

'<br>' +

event.message +

'</div>'

).join("");


}


<!-- RECENT TABLE -->


const table =
document.getElementById(
"recentTable"
);


const recent =
history.slice(-15).reverse();


table.innerHTML =
recent.map(
x =>

'<tr>' +

'<td>' +
x.time +
'</td>' +

'<td>' +
x.people +
'</td>' +

'<td>' +
x.temperature.toFixed(1) +
' °C</td>' +

'<td>' +
x.humidity.toFixed(1) +
' %</td>' +

'</tr>'

).join("");


}

catch(error) {

console.log(
"Dashboard error:",
error
);

}

}


<!-- ================================================= -->
<!-- EXPERIMENT MODE -->
<!-- ================================================= -->


function startExperiment() {


experimentStart =
Date.now();


document.getElementById(
"experimentResult"
).style.display =
"block";


document.getElementById(
"experimentResult"
).innerHTML =

"<strong>🧪 Experiment started.</strong><br>" +

"Collecting environmental data...";


}


async function finishExperiment() {


if (!experimentStart) {


alert(
"Start an experiment first."
);

return;

}


const response =
await fetch(
"/api/dashboard"
);


const data =
await response.json();


const live =
data.live;


const stats =
data.statistics;


const duration =
(
Date.now() -
experimentStart
) / 1000;


let result =

"<h3>Experiment Result</h3>" +

"<p><strong>Duration:</strong> " +
duration.toFixed(0) +
" seconds</p>";


if (live.online) {


result +=

"<p><strong>Current occupancy:</strong> " +
live.people +
"</p>" +

"<p><strong>Current temperature:</strong> " +
live.temperature.toFixed(1) +
" °C</p>" +

"<p><strong>Current humidity:</strong> " +
live.humidity.toFixed(1) +
" %</p>";

}


if (stats.reading_count) {


result +=

"<p><strong>Recorded average occupancy:</strong> " +
stats.average_people +
"</p>" +

"<p><strong>Recorded average temperature:</strong> " +
stats.average_temperature +
" °C</p>" +

"<p><strong>Recorded average humidity:</strong> " +
stats.average_humidity +
" %</p>";

}


result +=

"<p>Use the charts and event timeline to examine how the environment responded during the experiment.</p>";


document.getElementById(
"experimentResult"
).innerHTML =
result;


experimentStart = null;

}


<!-- ================================================= -->
<!-- START -->
<!-- ================================================= -->


makeCharts();

updateDashboard();

setInterval(
updateDashboard,
2000
);


</script>

</body>

</html>
"""


# =========================================================
# HOME
# =========================================================

@app.route("/")
def home():

    return render_template_string(HTML)


# =========================================================
# START SERVER
# =========================================================

if __name__ == "__main__":

    print("=================================")
    print(" INDUSTRIAL IoT INTELLIGENCE")
    print("=================================")
    print("Dashboard starting...")
    print()
    print("Open:")
    print("http://127.0.0.1:5000")
    print()

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=False
    )